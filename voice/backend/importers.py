from __future__ import annotations

import csv
import struct
from abc import ABC, abstractmethod
from dataclasses import dataclass
from itertools import islice
from pathlib import Path
from typing import Any, Iterator, Sequence

from openpyxl import load_workbook
from pyxlsb import biff12, open_workbook
import xlrd
from xlrd.compdoc import CompDoc

from .config import Settings, settings
from .errors import AppError
from .models import HEADER
from .security import detect_file_format


@dataclass(frozen=True, slots=True)
class RowRecord:
    source_row: int
    values: tuple[Any, ...]


class FileImporter(ABC):
    format: str

    def __init__(self, path: Path):
        self.path = Path(path)

    @classmethod
    def detect(
        cls, path: Path, original_name: str, config: Settings = settings
    ) -> str:
        return detect_file_format(Path(path), original_name, config)

    @abstractmethod
    def listSheets(self) -> list[str]:
        raise NotImplementedError

    def list_sheets(self) -> list[str]:
        return self.listSheets()

    def preview(self, sheet: str | None = None, limit: int = 20) -> list[RowRecord]:
        return list(islice(self.iterateRows(sheet), limit))

    @abstractmethod
    def iterateRows(self, sheet: str | None = None) -> Iterator[RowRecord]:
        raise NotImplementedError

    def iterate_rows(self, sheet: str | None = None) -> Iterator[RowRecord]:
        return self.iterateRows(sheet)

    def _resolve_sheet(self, requested: str | None) -> str:
        sheets = self.listSheets()
        if not sheets:
            raise AppError("EMPTY_FILE", "Файл не содержит листов")
        if requested is None:
            return sheets[0]
        if requested not in sheets:
            raise AppError("SHEET_NOT_FOUND", "Указанный лист не найден", status_code=400)
        return requested


class XlsxImporter(FileImporter):
    format = "xlsx"

    def listSheets(self) -> list[str]:
        try:
            # Uploads intentionally use opaque extensionless paths. Passing a
            # binary stream keeps openpyxl from rejecting a valid workbook
            # solely because that private storage path has no ".xlsx" suffix.
            with self.path.open("rb") as source:
                workbook = load_workbook(
                    source,
                    read_only=True,
                    data_only=False,
                    keep_links=False,
                )
                try:
                    return list(workbook.sheetnames)
                finally:
                    workbook.close()
        except Exception as exc:
            raise AppError("CORRUPT_FILE", "Не удалось прочитать XLSX") from exc

    def iterateRows(self, sheet: str | None = None) -> Iterator[RowRecord]:
        selected = self._resolve_sheet(sheet)
        try:
            with self.path.open("rb") as source:
                workbook = load_workbook(
                    source,
                    read_only=True,
                    data_only=False,
                    keep_links=False,
                )
                worksheet = workbook[selected]
                try:
                    for source_row, values in enumerate(
                        worksheet.iter_rows(values_only=True), start=1
                    ):
                        yield RowRecord(source_row, tuple(values))
                finally:
                    workbook.close()
        except AppError:
            raise
        except Exception as exc:
            raise AppError("CORRUPT_FILE", "Не удалось прочитать XLSX") from exc


class XlsImporter(FileImporter):
    format = "xls"

    def _reject_formula_records(self) -> None:
        try:
            compound = CompDoc(self.path.read_bytes())
            stream = compound.get_named_stream("Workbook")
            if stream is None:
                stream = compound.get_named_stream("Book")
            if stream is None:
                return
            position = 0
            while position + 4 <= len(stream):
                record_id, length = struct.unpack_from("<HH", stream, position)
                position += 4
                if position + length > len(stream):
                    break
                # BIFF2-BIFF8 FORMULA cell record. Rejecting the workbook is
                # deliberate: xlrd otherwise exposes only its cached result.
                if record_id == 0x0006:
                    raise AppError(
                        "FORMULA_CELL",
                        "Формулы в XLS не могут использоваться как номера",
                    )
                position += length
        except AppError:
            raise
        except Exception as exc:
            raise AppError("CORRUPT_FILE", "Не удалось проверить XLS") from exc

    def _open(self):
        try:
            self._reject_formula_records()
            return xlrd.open_workbook(self.path, on_demand=True)
        except AppError:
            raise
        except Exception as exc:
            raise AppError("CORRUPT_FILE", "Не удалось прочитать XLS") from exc

    def listSheets(self) -> list[str]:
        workbook = self._open()
        try:
            return workbook.sheet_names()
        finally:
            workbook.release_resources()

    def iterateRows(self, sheet: str | None = None) -> Iterator[RowRecord]:
        selected = self._resolve_sheet(sheet)
        workbook = self._open()
        try:
            worksheet = workbook.sheet_by_name(selected)
            for index in range(worksheet.nrows):
                yield RowRecord(index + 1, tuple(worksheet.row_values(index)))
        finally:
            workbook.release_resources()


class XlsbImporter(FileImporter):
    format = "xlsb"

    def listSheets(self) -> list[str]:
        try:
            with open_workbook(self.path) as workbook:
                return list(workbook.sheets)
        except Exception as exc:
            raise AppError("CORRUPT_FILE", "Не удалось прочитать XLSB") from exc

    def iterateRows(self, sheet: str | None = None) -> Iterator[RowRecord]:
        selected = self._resolve_sheet(sheet)
        try:
            with open_workbook(self.path) as workbook:
                with workbook.get_sheet(selected) as worksheet:
                    reader = worksheet._reader
                    reader.seek(worksheet._data_offset)
                    row_number = -1
                    values: list[Any] | None = None
                    width = worksheet.dimension.c + worksheet.dimension.w
                    formula_records = {
                        biff12.FORMULA_STRING,
                        biff12.FORMULA_FLOAT,
                        biff12.FORMULA_BOOL,
                        biff12.FORMULA_BOOLERR,
                    }
                    for record_id, record in reader:
                        if record_id == biff12.ROW and record.r != row_number:
                            if values is not None:
                                yield RowRecord(row_number + 1, tuple(values))
                            while row_number < record.r - 1:
                                row_number += 1
                                yield RowRecord(
                                    row_number + 1, tuple(None for _ in range(width))
                                )
                            row_number = record.r
                            values = [None] * width
                        elif (
                            biff12.BLANK <= record_id <= biff12.FORMULA_BOOLERR
                            and values is not None
                        ):
                            if record_id in formula_records:
                                value: Any = "=FORMULA"
                            elif (
                                record_id == biff12.STRING
                                and worksheet._stringtable is not None
                            ):
                                value = worksheet._stringtable[record.v]
                            else:
                                value = record.v
                            values[record.c] = value
                        elif record_id == biff12.SHEETDATA_END:
                            if values is not None:
                                yield RowRecord(row_number + 1, tuple(values))
                            break
        except AppError:
            raise
        except Exception as exc:
            raise AppError("CORRUPT_FILE", "Не удалось прочитать XLSB") from exc


class CsvImporter(FileImporter):
    format = "csv"
    sheet_name = "CSV"

    def _encoding(self) -> str:
        sample = self.path.read_bytes()[:256 * 1024]
        for encoding in ("utf-8-sig", "utf-8", "cp1251"):
            try:
                sample.decode(encoding)
                return encoding
            except UnicodeDecodeError:
                continue
        raise AppError("INVALID_ENCODING", "CSV должен быть в UTF-8 или Windows-1251")

    def _dialect(self, sample: str) -> csv.Dialect:
        try:
            return csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            return csv.excel

    @staticmethod
    def _formatted_one_column_delimiter(sample: str) -> str | None:
        # A one-column CSV has no physical separators, so a custom delimiter
        # cannot be inferred. The exact formatted header makes the grammar
        # unambiguous; an absent control character lets csv.reader safely
        # remove optional standard CSV quoting for any configured delimiter.
        for candidate in ("\x1f", "\x1e", "\t", "|", "^"):
            if candidate in sample:
                continue
            try:
                first = next(csv.reader(sample.splitlines()[:1], delimiter=candidate))
            except (csv.Error, StopIteration):
                continue
            if first == [HEADER]:
                return candidate
        return None

    def listSheets(self) -> list[str]:
        return [self.sheet_name]

    def iterateRows(self, sheet: str | None = None) -> Iterator[RowRecord]:
        self._resolve_sheet(sheet)
        encoding = self._encoding()
        try:
            with self.path.open("r", encoding=encoding, newline="") as source:
                sample = source.read(64 * 1024)
                source.seek(0)
                formatted_delimiter = self._formatted_one_column_delimiter(sample)
                reader = (
                    csv.reader(source, delimiter=formatted_delimiter)
                    if formatted_delimiter is not None
                    else csv.reader(source, self._dialect(sample))
                )
                for source_row, values in enumerate(reader, start=1):
                    yield RowRecord(source_row, tuple(values))
        except csv.Error as exc:
            raise AppError("CORRUPT_FILE", "Не удалось разобрать CSV") from exc


class FormattedMappingImporter(FileImporter):
    """A one-column view over another importer for formatted mapping rows."""

    format = "formatted"

    def __init__(self, delegate: FileImporter):
        super().__init__(delegate.path)
        self.delegate = delegate

    def listSheets(self) -> list[str]:
        return self.delegate.listSheets()

    def iterateRows(self, sheet: str | None = None) -> Iterator[RowRecord]:
        for row in self.delegate.iterateRows(sheet):
            first_non_empty = next(
                (value for value in row.values if value is not None and str(value).strip()),
                None,
            )
            yield RowRecord(row.source_row, (first_non_empty,))


def importer_for(path: Path, format_name: str) -> FileImporter:
    importer_types: dict[str, type[FileImporter]] = {
        "xlsx": XlsxImporter,
        "xls": XlsImporter,
        "xlsb": XlsbImporter,
        "csv": CsvImporter,
    }
    try:
        return importer_types[format_name](path)
    except KeyError as exc:
        raise AppError("UNSUPPORTED_FILE", "Неподдерживаемый формат файла") from exc


def row_value(values: Sequence[Any], index: int) -> Any:
    return values[index] if index < len(values) else None
