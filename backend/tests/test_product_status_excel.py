import io
import zipfile

from openpyxl import load_workbook

from app.product_status_excel import _unique_sheet_name, generate_b2b_product_status_excel
from app.schemas import ProductStatusB2BOut, ProductStatusSheetOut


def test_unique_sheet_name_sanitizes_invalid_chars() -> None:
    used: set[str] = set()
    assert _unique_sheet_name("Продуктовый офис: CORE", used) == "Продуктовый офис  CORE"
    assert _unique_sheet_name("Продуктовый офис: CORE", used) == "Продуктовый офис  CORE 1"


def test_generate_excel_workbook_contains_sheets_and_values() -> None:
    data = ProductStatusB2BOut(
        title="Статус продукта B2B",
        sheets=[
            ProductStatusSheetOut(
                gid="0",
                name="Продуктовый офис: CORE",
                columns=["Дата запуска", "Проект", "Описание проекта", "Зачем и для чего делаем"],
                rows=[
                    {
                        "Дата запуска": "09.06",
                        "Проект": "Ремонт",
                        "Описание проекта": "Тест",
                        "Зачем и для чего делаем": "OK",
                    }
                ],
                totalShown=1,
            ),
            ProductStatusSheetOut(
                gid="1",
                name="Продуктовый офис: SMS",
                columns=["Дата запуска", "Проект", "Описание проекта", "Зачем и для чего делаем"],
                rows=[
                    {
                        "Дата запуска": "",
                        "Проект": "SMS-Таргет",
                        "Описание проекта": "Статус",
                        "Зачем и для чего делаем": "",
                    }
                ],
                totalShown=1,
            ),
        ],
    )

    content, filename = generate_b2b_product_status_excel(data)
    assert filename.endswith(".xlsx")
    assert content[:2] == b"PK"

    with zipfile.ZipFile(io.BytesIO(content)) as archive:
        assert any(name.startswith("xl/worksheets/") for name in archive.namelist())

    workbook = load_workbook(io.BytesIO(content))
    assert len(workbook.sheetnames) == 2
    core = workbook["Продуктовый офис  CORE"]
    assert core["A1"].value == "Дата запуска"
    assert core["B2"].value == "Ремонт"
    assert core["C2"].value == "Тест"


def test_generate_excel_applies_yellow_fill_for_highlighted_cells() -> None:
    data = ProductStatusB2BOut(
        title="Статус продукта B2B",
        sheets=[
            ProductStatusSheetOut(
                gid="0",
                name="CORE",
                columns=["Описание"],
                rows=[{"Описание": "Убираем $300 рублевые$ офферы"}],
                totalShown=1,
            )
        ],
    )

    content, _ = generate_b2b_product_status_excel(data)
    workbook = load_workbook(io.BytesIO(content))
    cell = workbook["CORE"]["A2"]
    assert cell.value == "Убираем 300 рублевые офферы"
    assert cell.fill.fgColor.rgb in {"00FFFF00", "FFFF00"}
