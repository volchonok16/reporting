import io

from openpyxl import load_workbook

from app.boards import get_boards
from app.report_service import export_xlsx
from app.schemas import ChangeRequestOut, DashboardMetricsOut, DashboardOut, LinkedErrorOut, LinkedProductOut


def _empty_metrics() -> DashboardMetricsOut:
    return DashboardMetricsOut(
        totalTasks=0,
        inProgress=0,
        launchingSoon=0,
        launched=0,
        completed=0,
        errorsCount=0,
    )


def test_export_xlsx_is_real_excel_file(monkeypatch) -> None:
    monkeypatch.setattr(
        "app.report_service.ensure_boards_loaded",
        lambda db, refresh=False: get_boards(),
    )

    def fake_load(_db, *, board_code=None, **_kwargs):
        items: list[ChangeRequestOut] = []
        if board_code == "b2b_product_core":
            items = [
                ChangeRequestOut(
                    id="1",
                    number="123456",
                    title="Тест ЗНИ",
                    customerName="Иванов Иван",
                    product=LinkedProductOut(
                        id="100",
                        title="Продукт A",
                        url="https://tfs.example/100",
                    ),
                    status="Active",
                    boardColumn="Development",
                    boardName="CORE",
                    errors=[LinkedErrorOut(id="9", title="ошибка")],
                ),
                ChangeRequestOut(
                    id="2",
                    number="222",
                    title="[EFO] скрыть",
                    boardName="CORE",
                ),
            ]
        return DashboardOut(metrics=_empty_metrics(), items=items, totalShown=len(items))

    monkeypatch.setattr("app.report_service.load_change_requests", fake_load)

    content, filename = export_xlsx(object(), board_code="b2b_product_core")
    assert filename.startswith("zni-report-b2b_product_core-")
    assert filename.endswith(".xlsx")
    assert content[:2] == b"PK"

    workbook = load_workbook(io.BytesIO(content))
    rows = list(workbook.active.iter_rows(values_only=True))
    assert rows[0][0] == "Номер ЗНИ"
    assert rows[0][2] == "Заказчик"
    assert rows[0][3] == "Продукт"
    assert rows[1][0] == "123456"
    assert rows[1][1] == "Тест ЗНИ"
    assert rows[1][2] == "Иванов Иван"
    assert rows[1][3] == "100: Продукт A"
    assert rows[1][13] == "9: ошибка"
    assert workbook.active.cell(row=2, column=4).hyperlink.target == "https://tfs.example/100"
    assert all(row[0] != "222" for row in rows[1:])


def test_export_all_collapses_same_alias_boards(monkeypatch) -> None:
    from app.boards import BoardConfig, set_boards_cache
    from app.report_service import _boards_for_export

    extra = BoardConfig(
        code="products_other",
        name="Другие продукты",
        display_name="Продукты",
        project="Tele2",
        project_id="pid",
        team_id="tid",
        area_path=r"Tele2\Other\Products",
    )
    boards = [*get_boards(), extra]
    set_boards_cache(boards)
    exported = _boards_for_export("all", boards)
    assert [board.code for board in exported].count("tele2_products") == 1
    assert all(board.code != "products_other" for board in exported)
    assert {board.code for board in _boards_for_export("products_other", boards)} == {"tele2_products"}
