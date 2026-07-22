from __future__ import annotations

import io

from openpyxl import load_workbook

from app.revenue_activities_excel import generate_revenue_activities_excel
from app.schemas import ProductStatusB2BOut, ProductStatusSheetOut


def test_revenue_excel_exports_numbers_as_numeric_cells() -> None:
    data = ProductStatusB2BOut(
        title="Активности по выручкам",
        sheets=[
            ProductStatusSheetOut(
                gid="base",
                name="Влияние по базе",
                columns=[
                    "Активность",
                    "Статус F2 2026",
                    "Ответственный",
                    "Влияние на базу, тыс",
                    "Влияние на gmc, млн",
                    "Комментарий",
                ],
                rows=[
                    {
                        "Активность": "Пилот",
                        "Статус F2 2026": "В работе",
                        "Ответственный": "Иванов",
                        "Влияние на базу, тыс": "10",
                        "Влияние на gmc, млн": "1",
                        "Комментарий": "ок",
                    }
                ],
                totalShown=1,
            ),
            ProductStatusSheetOut(
                gid="revenue",
                name="Влияние по выручке",
                columns=[
                    "Активность",
                    "Статус F2 2026",
                    "Ответственный",
                    "Влияние на выручку, млн",
                    "Маржа",
                    "Влияние на gmc, млн",
                    "Комментарий",
                ],
                rows=[
                    {
                        "Активность": "Пилот",
                        "Статус F2 2026": "В работе",
                        "Ответственный": "Иванов",
                        "Влияние на выручку, млн": "2,5",
                        "Маржа": "0,4",
                        "Влияние на gmc, млн": "1",
                        "Комментарий": "ок",
                    }
                ],
                totalShown=1,
            ),
        ],
    )

    content, filename = generate_revenue_activities_excel(data)
    assert filename.startswith("aktivnosti-po-vyruchkam-")
    assert filename.endswith(".xlsx")

    workbook = load_workbook(io.BytesIO(content))
    base_sheet = workbook["Влияние по базе"]
    assert base_sheet["A1"].value == "Активность"
    assert base_sheet["B1"].value == "Статус F2 2026"
    assert base_sheet["D1"].value == "Влияние на базу, тыс"
    assert base_sheet["A2"].value == "Пилот"
    assert base_sheet["D2"].value == 10
    assert base_sheet["E2"].value == 1
    assert base_sheet["F2"].value == "ок"
    assert base_sheet["A3"].value == "Итого"
    assert base_sheet["D3"].value == 10
    assert base_sheet["E3"].value == 1

    revenue_sheet = workbook["Влияние по выручке"]
    assert revenue_sheet["D1"].value == "Влияние на выручку, млн"
    assert revenue_sheet["E1"].value == "Маржа"
    assert revenue_sheet["D2"].value == 2.5
    assert revenue_sheet["E2"].value == 0.4
    assert revenue_sheet["F2"].value == 1
    assert revenue_sheet["D3"].value == 2.5
    assert revenue_sheet["E3"].value == 0.4


def test_revenue_excel_keeps_non_numeric_influence_as_text() -> None:
    data = ProductStatusB2BOut(
        title="Активности по выручкам",
        sheets=[
            ProductStatusSheetOut(
                gid="base",
                name="Влияние по базе",
                columns=["Активность", "Влияние на базу, тыс", "Комментарий"],
                rows=[
                    {
                        "Активность": "X",
                        "Влияние на базу, тыс": "н/д",
                        "Комментарий": "",
                    }
                ],
                totalShown=1,
            )
        ],
    )
    content, _ = generate_revenue_activities_excel(data)
    workbook = load_workbook(io.BytesIO(content))
    sheet = workbook.active
    assert sheet["B2"].value == "н/д"
