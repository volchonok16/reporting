from __future__ import annotations

import io
from decimal import Decimal

from openpyxl import load_workbook

from app.org_schemas import (
    DepartmentOut,
    EmployeeDepartmentBriefOut,
    EmployeeExpertiseOut,
    EmployeeOut,
)
from app.org_staffing_excel import generate_staffing_excel


def _employee(
    *,
    emp_id: int,
    name: str,
    departments: list[EmployeeDepartmentBriefOut],
    is_active: bool = True,
    hide_from_pyramid: bool = False,
) -> EmployeeOut:
    return EmployeeOut(
        id=emp_id,
        publicId=f"00000000-0000-0000-0000-{emp_id:012d}",
        fullName=name,
        email=f"{name.split()[0].lower()}@example.com",
        position="Инженер",
        managerName="Руководитель",
        dailyWorkHours=Decimal("8"),
        isActive=is_active,
        isOrganizationHead=False,
        hideFromPyramid=hide_from_pyramid,
        expertises=[
            EmployeeExpertiseOut(id=1, directionId=1, directionName="Backend", level="Senior"),
        ],
        departments=departments,
    )


def test_staffing_excel_exports_pyramid_people_without_director_column() -> None:
    departments = [
        DepartmentOut(
            id=10,
            name="Платформа",
            sortOrder=1,
            isActive=True,
            memberCount=2,
        ),
        DepartmentOut(
            id=20,
            name="Пустой отдел",
            sortOrder=2,
            isActive=True,
            memberCount=0,
        ),
        DepartmentOut(
            id=30,
            name="Аналитика",
            sortOrder=0,
            isActive=True,
            memberCount=1,
        ),
    ]
    employees = [
        _employee(
            emp_id=1,
            name="Белов Борис",
            departments=[EmployeeDepartmentBriefOut(departmentId=10, departmentName="Платформа")],
        ),
        _employee(
            emp_id=2,
            name="Алова Анна",
            departments=[
                EmployeeDepartmentBriefOut(departmentId=10, departmentName="Платформа"),
                EmployeeDepartmentBriefOut(departmentId=30, departmentName="Аналитика"),
            ],
        ),
        _employee(
            emp_id=3,
            name="Сидоров Сидор",
            departments=[],
        ),
        _employee(
            emp_id=4,
            name="Скрытый Сергей",
            departments=[EmployeeDepartmentBriefOut(departmentId=10, departmentName="Платформа")],
            hide_from_pyramid=True,
        ),
        _employee(
            emp_id=5,
            name="Неактивный Николай",
            departments=[EmployeeDepartmentBriefOut(departmentId=30, departmentName="Аналитика")],
            is_active=False,
        ),
    ]

    content, filename = generate_staffing_excel(employees, departments)
    assert filename.endswith(".xlsx")

    workbook = load_workbook(filename=io.BytesIO(content))
    assert workbook.sheetnames == ["Общий список", "Аналитика", "Платформа"]

    general = workbook["Общий список"]
    assert [general.cell(1, col).value for col in range(1, 10)] == [
        "ФИО",
        "Должность",
        "Отделы",
        "Email",
        "Рабочих часов в день",
        "Экспертиза",
        "Руководитель",
        "Активен",
        None,
    ]
    assert general.cell(2, 1).value == "Алова Анна"
    assert general.cell(3, 1).value == "Белов Борис"
    assert general.cell(4, 1).value == "Сидоров Сидор"
    assert general.max_row == 4

    analytics = workbook["Аналитика"]
    assert analytics.cell(1, 8).value is None
    assert analytics.cell(2, 1).value == "Алова Анна"
    assert analytics.max_row == 2

    platform = workbook["Платформа"]
    assert platform.cell(2, 1).value == "Алова Анна"
    assert platform.cell(3, 1).value == "Белов Борис"
    assert platform.max_row == 3
