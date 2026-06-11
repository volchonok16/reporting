from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.product_status_xlsx_source import _cell_encoded_value, _parse_xlsx_sheet


def test_cell_encoded_value_applies_foreground_and_strike() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    cell = worksheet["A1"]
    cell.value = "Индексация 2027"
    cell.font = Font(color="FFFF0000", strike=True)

    encoded = _cell_encoded_value(cell)
    assert encoded == "[[fg:FF0000;strike::Индексация 2027]]"


def test_cell_encoded_value_applies_cell_background() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    cell = worksheet["A1"]
    cell.value = "Идет миграция"
    cell.fill = PatternFill(fill_type="solid", fgColor="FFC6EFCE")

    encoded = _cell_encoded_value(cell)
    assert encoded.startswith("<<cell:bg:C6EFCE>>")
    assert "Идет миграция" in encoded


def test_parse_xlsx_sheet_reads_headers_and_rows() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["A1"] = "Дата запуска"
    worksheet["B1"] = "Проект"
    worksheet["A2"] = "09.06"
    worksheet["B2"] = "Ремонт"

    columns, rows = _parse_xlsx_sheet(worksheet)
    assert columns == ["Дата запуска", "Проект"]
    assert rows == [{"Дата запуска": "09.06", "Проект": "Ремонт"}]
