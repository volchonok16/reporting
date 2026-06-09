from __future__ import annotations

import io
import re
from datetime import datetime
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.product_status_rich_text import (
    cell_highlight_colors,
    display_cell_text,
    split_cell_wrapper,
    split_style_segments,
)
from app.schemas import ProductStatusB2BOut, ProductStatusSheetOut

MOSCOW_TZ = ZoneInfo("Europe/Moscow")
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="CCCCCC")
_HEADER_FONT = Font(bold=True)
_CELL_ALIGNMENT = Alignment(wrap_text=True, vertical="top")

def _highlight_fill(color_hex: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=color_hex.upper())
_INVALID_SHEET_CHARS = re.compile(r"[\:\\/?*\[\]]+")
_MAX_SHEET_NAME_LEN = 31


def _unique_sheet_name(raw: str, used: set[str]) -> str:
    cleaned = _INVALID_SHEET_CHARS.sub(" ", (raw or "Лист").strip()) or "Лист"
    base = cleaned[:_MAX_SHEET_NAME_LEN].strip() or "Лист"
    candidate = base
    suffix = 1
    while candidate.casefold() in used:
        tail = f" {suffix}"
        candidate = f"{base[: _MAX_SHEET_NAME_LEN - len(tail)]}{tail}".strip()
        suffix += 1
    used.add(candidate.casefold())
    return candidate


def _autosize_columns(
    worksheet: Worksheet,
    *,
    column_count: int,
    row_count: int,
) -> None:
    for index in range(1, column_count + 1):
        letter = get_column_letter(index)
        max_len = 12
        for row_index in range(1, row_count + 1):
            value = worksheet.cell(row=row_index, column=index).value
            if value is None:
                continue
            lines = str(value).splitlines() or [""]
            max_len = max(max_len, max(len(line) for line in lines))
        width = min(72, max(12, max_len + 2))
        worksheet.column_dimensions[letter].width = width


def _write_sheet(worksheet: Worksheet, sheet: ProductStatusSheetOut) -> None:
    columns = sheet.columns
    for col_index, column_name in enumerate(columns, start=1):
        cell = worksheet.cell(row=1, column=col_index, value=column_name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CELL_ALIGNMENT

    for row_index, row in enumerate(sheet.rows, start=2):
        for col_index, column_name in enumerate(columns, start=1):
            raw = row.get(column_name, "") if column_name else ""
            value = display_cell_text(raw).strip()
            cell = worksheet.cell(row=row_index, column=col_index, value=value or None)
            cell.alignment = _CELL_ALIGNMENT
            cell_style, inner = split_cell_wrapper(raw)
            if cell_style.bg:
                cell.fill = _highlight_fill(cell_style.bg)
            elif cell_highlight_colors(raw):
                cell.fill = _highlight_fill(cell_highlight_colors(raw)[0])
            if cell_style.border:
                side = Side(style="thin", color=cell_style.border)
                cell.border = Border(left=side, right=side, top=side, bottom=side)
            segments = split_style_segments(inner)
            first = next((segment for segment in segments if segment.text.strip()), None)
            if first and (first.fg or first.strike or first.bold or first.italic):
                cell.font = Font(
                    color=first.fg or None,
                    strike=first.strike,
                    bold=first.bold,
                    italic=first.italic,
                )

    _autosize_columns(
        worksheet,
        column_count=len(columns),
        row_count=max(1, len(sheet.rows) + 1),
    )
    worksheet.freeze_panes = "A2"


def generate_b2b_product_status_excel(data: ProductStatusB2BOut) -> tuple[bytes, str]:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    used_names: set[str] = set()
    for sheet in data.sheets:
        title = _unique_sheet_name(sheet.name, used_names)
        worksheet = workbook.create_sheet(title=title)
        _write_sheet(worksheet, sheet)

    if not workbook.sheetnames:
        worksheet = workbook.create_sheet(title="Статус")
        _write_sheet(
            worksheet,
            ProductStatusSheetOut(
                gid="0",
                name="Статус",
                columns=[],
                rows=[],
                totalShown=0,
            ),
        )

    buffer = io.BytesIO()
    workbook.save(buffer)
    generated_at = datetime.now(MOSCOW_TZ)
    stamp = generated_at.strftime("%Y%m%d")
    filename = f"status-produkta-b2b-{stamp}.xlsx"
    return buffer.getvalue(), filename
