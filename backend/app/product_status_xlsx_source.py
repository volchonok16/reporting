from __future__ import annotations

import io
import logging
import time

import httpx
from openpyxl import load_workbook
from app.product_status_rich_text import (
    _filter_cell_background,
    encode_style_segment,
    wrap_cell_text,
)

logger = logging.getLogger(__name__)

_DEFAULT_TEXT_COLORS = frozenset({"172B4D", "000000"})
_DEFAULT_FILL_COLORS = frozenset({"FFFFFF", "000000"})


def _xlsx_export_url(spreadsheet_id: str, gid: str) -> str:
    cache_bust = int(time.time())
    return (
        "https://docs.google.com/spreadsheets/d/"
        f"{spreadsheet_id}/export?format=xlsx&gid={gid}&_={cache_bust}"
    )


def _openpyxl_color_to_hex(color) -> str | None:
    if color is None or getattr(color, "type", None) != "rgb" or not color.rgb:
        return None
    value = str(color.rgb).upper().lstrip("#")
    if len(value) == 8:
        value = value[2:]
    if len(value) != 6:
        return None
    if value in _DEFAULT_TEXT_COLORS or value in _DEFAULT_FILL_COLORS:
        return None
    return value


def _border_color(cell) -> str | None:
    # XLSX-экспорт Google Sheets часто добавляет серую сетку на все ячейки — не переносим.
    return None


def _cell_plain_value(cell) -> str:
    value = cell.value
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        try:
            return value.strftime("%d.%m.%Y")
        except (TypeError, ValueError):
            pass
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def _cell_encoded_value(cell) -> str:
    text = _cell_plain_value(cell)
    if not text:
        return ""

    font = cell.font
    fill = cell.fill
    fg = _openpyxl_color_to_hex(font.color) if font else None
    cell_bg = None
    if fill and getattr(fill, "fill_type", None) == "solid":
        cell_bg = _openpyxl_color_to_hex(fill.fgColor)
    border = _border_color(cell)

    styled = text
    if fg or (font and (font.strike or font.bold or font.italic)):
        styled = encode_style_segment(
            text,
            fg=fg,
            strike=bool(font and font.strike),
            bold=bool(font and font.bold),
            italic=bool(font and font.italic),
        )

    cell_bg = _filter_cell_background(cell_bg, styled)
    return wrap_cell_text(styled, bg=cell_bg, border=border)


def _parse_xlsx_sheet(worksheet) -> tuple[list[str], list[dict[str, str]]]:
    rows_iter = worksheet.iter_rows(values_only=False)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], []

    headers = [_cell_plain_value(cell) for cell in header_row]
    headers = [header for header in headers if header]
    if not headers:
        return [], []

    rows: list[dict[str, str]] = []
    for raw_row in rows_iter:
        row_values: dict[str, str] = {}
        has_content = False
        for index, header in enumerate(headers):
            cell = raw_row[index] if index < len(raw_row) else None
            value = _cell_encoded_value(cell) if cell is not None else ""
            row_values[header] = value
            if value:
                has_content = True
        if has_content:
            rows.append(row_values)
    return headers, rows


def fetch_sheet_from_xlsx(
    *,
    spreadsheet_id: str,
    gid: str,
    client: httpx.Client,
) -> tuple[list[str], list[dict[str, str]]] | None:
    url = _xlsx_export_url(spreadsheet_id, gid)
    try:
        response = client.get(url)
        response.raise_for_status()
    except httpx.HTTPError:
        logger.warning(
            "product_status_xlsx_fetch_failed gid=%s",
            gid,
            exc_info=True,
        )
        return None

    try:
        workbook = load_workbook(io.BytesIO(response.content), data_only=False)
    except Exception:
        logger.warning("product_status_xlsx_parse_failed gid=%s", gid, exc_info=True)
        return None

    worksheet = workbook.active
    if worksheet is None:
        return None
    return _parse_xlsx_sheet(worksheet)
