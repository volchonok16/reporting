import io
from datetime import date, datetime, timedelta
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import or_, select
from sqlalchemy.orm import Session

from app.boards import (
    BoardConfig,
    boards_for_sync,
    boards_sharing_alias,
    ensure_boards_loaded,
    get_boards,
    grouped_boards_for_ui,
    is_all_boards,
)
from app.board_metrics import (
    active_errors,
    board_for_task,
    has_linked_errors,
    is_completed,
    is_in_progress,
    is_launched,
    is_launching_soon,
    task_status_tokens,
)
from app.config import settings
from app.iteration_plan import (
    PLAN_QUARTER_TBD,
    parse_iteration_plan,
    quarter_key_from_date,
    quarter_label_from_key,
)
from app.models import Task, ZniExternalData
from app.roadmap_priority_service import roadmap_comment_from_task, roadmap_priority_from_task
from app.digital_plan_service import ect_acceptance_from_task, has_uc_from_task
from app.zni_external_data_service import actual_period_editable_statuses, load_external_data_by_task_ids
from app.completed_metrics import has_customer_name
from app.resource_reservation import ect_resource_reservation_label
from app.zni_description import tfs_identity_display_name
from app.schemas import (
    BoardOut,
    ChangeRequestOut,
    DashboardMetricsOut,
    DashboardOut,
    LinkedEnvironmentOut,
    LinkedErrorOut,
    LinkedProductOut,
    QuarterOptionOut,
    TagFilterGroupOut,
)
from app.zni_linked_environments import has_linked_environment, linked_environment_records_from_extra
from app.tag_filters import (
    DIGITAL_BOARD_CODE,
    normalize_tag_group_keys,
    tag_filter_groups_for_board,
    tag_filter_supported_for_board,
    task_matches_tag_groups,
)
from app.zni_title_filters import is_excluded_zni_title

BERCUT_BOARD_CODE = "be_t2_team"
INCIDENT_ERROR_ROW_TYPE = "error"


def _board_out(board: BoardConfig, boards: list[BoardConfig] | None = None) -> BoardOut:
    members = [item.code for item in boards_sharing_alias(board, boards)]
    return BoardOut(
        code=board.code,
        name=board.name,
        displayName=board.display_name,
        project=board.project,
        memberCodes=members,
    )


def _board_task_clause(board: BoardConfig):
    names = {name for name in (board.name, board.display_name) if name}
    return or_(
        Task.extra_json["board_code"].as_string() == board.code,
        Task.source_team.in_(sorted(names)),
    )


def _boards_task_clause(boards: list[BoardConfig]):
    clauses = [_board_task_clause(board) for board in boards]
    if not clauses:
        return False
    if len(clauses) == 1:
        return clauses[0]
    return or_(*clauses)


def _empty_metrics() -> DashboardMetricsOut:
    return DashboardMetricsOut(
        totalTasks=0,
        inProgress=0,
        launchingSoon=0,
        launched=0,
        completed=0,
        errorsCount=0,
    )


def _matches_search(task: Task, search: str) -> bool:
    if not search:
        return True
    needle = search.strip().lower()
    return needle in task.external_id.lower() or needle in task.title.lower()


def _extra(task: Task) -> dict:
    return task.extra_json if isinstance(task.extra_json, dict) else {}


def _task_tags(task: Task) -> list[str]:
    raw = _extra(task).get("tags")
    if not isinstance(raw, list):
        return []
    return [str(tag).strip() for tag in raw if str(tag).strip()]


def _matches_tag_groups(task: Task, tag_groups: list[str]) -> bool:
    return task_matches_tag_groups(_task_tags(task), tag_groups)


def _tag_filter_groups_out(board_code: str | None) -> list[TagFilterGroupOut]:
    return [
        TagFilterGroupOut(
            key=group.key,
            label=group.label,
            tags=list(group.root_tags),
            subsectionPrefixes=list(group.subsection_prefixes),
        )
        for group in tag_filter_groups_for_board(board_code)
    ]


def _planned_release(task: Task) -> str | None:
    value = _extra(task).get("planned_release")
    return str(value).strip() if value else None


def _ect_resource_reservation(task: Task) -> bool:
    value = _extra(task).get("ect_resource_reservation")
    return value is True


def _linked_environments(task: Task) -> list[LinkedEnvironmentOut]:
    items: list[LinkedEnvironmentOut] = []
    for record in linked_environment_records_from_extra(_extra(task)):
        zni_id = str(record.get("zni_id") or record.get("zniId") or "").strip()
        if not zni_id:
            continue
        key = str(record.get("key") or "").strip()
        label = str(record.get("label") or key or "").strip() or key
        status = record.get("status")
        board_column = record.get("board_column") or record.get("boardColumn")
        url = record.get("url")
        items.append(
            LinkedEnvironmentOut(
                key=key,
                label=label,
                zniId=zni_id,
                status=str(status).strip() if status not in (None, "") else None,
                boardColumn=str(board_column).strip() if board_column not in (None, "") else None,
                url=str(url).strip() if url not in (None, "") else None,
            )
        )
    return items


def _customer_name(task: Task) -> str | None:
    value = _extra(task).get("customer_name")
    if not value:
        return None
    return tfs_identity_display_name(value)


def _business_goal(task: Task) -> str | None:
    value = _extra(task).get("business_goal")
    return str(value).strip() if value else None


def _business_value(task: Task) -> int | None:
    value = _extra(task).get("business_value")
    if value is None or value == "":
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _number_tiebreaker(task: Task) -> int:
    try:
        return int(task.external_id)
    except ValueError:
        return 0


def _business_value_asc_sort_key(task: Task) -> tuple[int, int, int]:
    """1 → больше; без значения — в конце."""
    value = _business_value(task)
    if value is None:
        return (1, 0, _number_tiebreaker(task))
    return (0, value, _number_tiebreaker(task))


def _business_value_desc_sort_key(task: Task) -> tuple[int, int, int]:
    """Без значения — в начале; затем от большего к меньшему."""
    value = _business_value(task)
    if value is None:
        return (0, 0, _number_tiebreaker(task))
    return (1, -value, _number_tiebreaker(task))


def _board_column(task: Task) -> str | None:
    value = _extra(task).get("board_column")
    return str(value).strip() if value else None


def _task_plan_meta(task: Task) -> tuple[date | None, str | None, str | None, str | None]:
    extra = _extra(task)
    planned_raw = extra.get("planned_date")
    quarter_key = extra.get("plan_quarter")
    planned_status = extra.get("planned_status")
    planned: date | None = None
    if isinstance(planned_raw, str) and planned_raw:
        try:
            planned = date.fromisoformat(planned_raw)
        except ValueError:
            planned = None

    is_tbd = planned_status == "tbd" or quarter_key == PLAN_QUARTER_TBD
    if not is_tbd and planned is None:
        iteration_path = extra.get("iteration_path") or task.iteration_path
        if isinstance(iteration_path, str):
            plan = parse_iteration_plan(iteration_path)
            is_tbd = plan.is_tbd
            planned = plan.planned_date
            if not quarter_key:
                quarter_key = plan.quarter_key

    if not is_tbd and planned is None and task.release_date:
        planned = task.release_date

    if is_tbd:
        return None, PLAN_QUARTER_TBD, "TBD", "TBD"

    if planned and not quarter_key:
        quarter_key = quarter_key_from_date(planned)
    quarter_key_str = str(quarter_key).strip() if quarter_key else None
    quarter_label = quarter_label_from_key(quarter_key_str) if quarter_key_str else None
    return planned, quarter_key_str, quarter_label, None


def _matches_quarter(task: Task, quarter: str | None) -> bool:
    if not quarter:
        return True
    _, quarter_key, _, _ = _task_plan_meta(task)
    if quarter == "__none__":
        return quarter_key is None
    if quarter == PLAN_QUARTER_TBD:
        return quarter_key == PLAN_QUARTER_TBD
    return quarter_key == quarter


def _matches_ect_reservation(task: Task, ect_reservation: str | None) -> bool:
    if not ect_reservation:
        return True
    has = _ect_resource_reservation(task)
    if ect_reservation == "yes":
        return has
    if ect_reservation == "no":
        return not has
    return True


def _matches_linked_environment(task: Task, linked_environment: str | None) -> bool:
    if linked_environment != "yes":
        return True
    return has_linked_environment(_extra(task))


def _matches_status(task: Task, status: str | None) -> bool:
    if not status:
        return True
    needle = status.strip().lower()
    column = (_board_column(task) or "").lower()
    workflow = (task.source_status or "").lower()
    return needle in {column, workflow}


def _is_incident_standalone_error(task: Task) -> bool:
    if task.task_type != "error" or task.parent_task_id is not None:
        return False
    return _extra(task).get("incident_error") is True


def _standalone_incident_errors(error_rows: list[Task]) -> list[Task]:
    return [row for row in error_rows if _is_incident_standalone_error(row)]


def _incident_error_visible_for_metric(metric: str | None) -> bool:
    return not metric or metric == "errors"


def _matches_incident_error_row(
    error: Task,
    *,
    board_code: str | None,
    metric: str | None,
    search: str | None,
    status: str | None,
    date_from: date | None,
    date_to: date | None,
) -> bool:
    if not _is_incident_standalone_error(error):
        return False
    error_board_code = str(_extra(error).get("board_code") or "").strip().lower()
    if error_board_code != BERCUT_BOARD_CODE:
        return False
    if board_code and not is_all_boards(board_code):
        if board_code.strip().lower() != BERCUT_BOARD_CODE:
            return False
    if not _incident_error_visible_for_metric(metric):
        return False
    if not _matches_search(error, search or ""):
        return False
    if not _matches_status(error, status):
        return False
    if not _in_date_range(error, date_from, date_to):
        return False
    return True


def _product_out(product: Task | None) -> LinkedProductOut | None:
    if product is None:
        return None
    return LinkedProductOut(
        id=product.external_id,
        title=product.title,
        url=product.external_url,
    )


def _load_products_by_id(db: Session, rows: list[Task]) -> dict[int, Task]:
    product_ids = {row.parent_task_id for row in rows if row.parent_task_id is not None}
    if not product_ids:
        return {}
    products = list(
        db.scalars(
            select(Task).where(
                Task.id.in_(product_ids),
                Task.task_type == "product",
            )
        )
    )
    return {product.id: product for product in products}


def _change_request_to_out(
    row: Task,
    linked_errors: list[Task],
    external: ZniExternalData | None = None,
    product: Task | None = None,
) -> ChangeRequestOut:
    board_code_value = _extra(row).get("board_code")
    planned_date, _, quarter_label, planned_label = _task_plan_meta(row)
    return ChangeRequestOut(
        id=str(row.id),
        number=row.external_id,
        rowType="change_request",
        title=row.title,
        url=row.external_url,
        status=row.source_status,
        boardColumn=_board_column(row),
        startDate=_effective_start(row),
        releaseDate=row.release_date,
        plannedDate=planned_date,
        plannedLabel=planned_label,
        planQuarter=quarter_label,
        plannedRelease=_planned_release(row),
        createdAt=row.created_at,
        boardCode=str(board_code_value) if board_code_value else None,
        boardName=row.source_team or _board_name_by_code(str(board_code_value) if board_code_value else None),
        customerName=_customer_name(row),
        product=_product_out(product),
        businessGoal=_business_goal(row),
        businessValue=_business_value(row),
        roadmapPriority=roadmap_priority_from_task(row),
        roadmapComment=roadmap_comment_from_task(row),
        ectResourceReservation=_ect_resource_reservation(row),
        ectAcceptance=ect_acceptance_from_task(row),
        hasUc=has_uc_from_task(row),
        linkedEnvironments=_linked_environments(row),
        errors=[
            LinkedErrorOut(
                id=error.external_id,
                title=error.title,
                status=error.source_status,
                url=error.external_url,
            )
            for error in linked_errors
        ],
        externalPriority=external.priority if external else None,
        externalCommercialEffect=external.commercial_effect if external else None,
        externalActualPeriod=external.actual_period if external else None,
        externalDesiredDate=external.desired_date if external else None,
        externalComment=external.comment if external else None,
        externalCategoryId=external.category_id if external else None,
        externalCategoryName=(
            external.category.name if external and external.category is not None else None
        ),
    )


def load_change_requests_by_numbers(db: Session, numbers: list[str]) -> list[ChangeRequestOut]:
    normalized: list[str] = []
    seen: set[str] = set()
    for raw in numbers:
        number = raw.strip()
        if not number or not number.isdigit() or number in seen:
            continue
        seen.add(number)
        normalized.append(number)
    if not normalized:
        return []

    boards = ensure_boards_loaded(db)
    scope = _boards_task_clause(boards)
    rows = list(
        db.scalars(
            select(Task).where(
                Task.task_type == "change_request",
                Task.external_id.in_(normalized),
                scope,
            )
        )
    )
    if not rows:
        return []

    error_rows = active_errors(
        list(
            db.scalars(
                select(Task).where(
                    Task.task_type == "error",
                    scope,
                )
            )
        )
    )
    errors_by_parent = _build_errors_by_parent(rows, error_rows)
    external_by_id = load_external_data_by_task_ids(db, [row.id for row in rows])
    products_by_id = _load_products_by_id(db, rows)
    by_number = {
        row.external_id: _change_request_to_out(
            row,
            errors_by_parent.get(row.id, []),
            external_by_id.get(row.id),
            products_by_id.get(row.parent_task_id) if row.parent_task_id else None,
        )
        for row in rows
    }
    return [by_number[number] for number in normalized if number in by_number]


def _incident_error_to_item(error: Task) -> ChangeRequestOut:
    board_code_value = _extra(error).get("board_code")
    return ChangeRequestOut(
        id=str(error.id),
        number=error.external_id,
        rowType=INCIDENT_ERROR_ROW_TYPE,
        title=error.title,
        url=error.external_url,
        status=error.source_status,
        boardColumn=_board_column(error),
        startDate=_effective_start(error),
        releaseDate=error.release_date,
        createdAt=error.created_at,
        boardCode=str(board_code_value) if board_code_value else None,
        boardName=error.source_team or _board_name_by_code(str(board_code_value) if board_code_value else None),
        errors=[],
    )


def _effective_start(task: Task) -> date | None:
    return task.start_date or (task.created_at.date() if task.created_at else None)


def _in_date_range(task: Task, date_from: date | None, date_to: date | None) -> bool:
    if not date_from and not date_to:
        return True
    start = _effective_start(task)
    if start is None:
        return False
    if date_from and start < date_from:
        return False
    if date_to and start > date_to:
        return False
    return True


def _row_board_code(row: Task, dashboard_board_code: str | None) -> str | None:
    """На «Все доски» правила метрик — по доске задачи, не по фильтру дашборда."""
    if is_all_boards(dashboard_board_code):
        board = board_for_task(row)
        return board.code if board else None
    return dashboard_board_code


def _board_metrics_ignore_date_period(board_code: str | None) -> bool:
    """Digital и Bercut: карточки метрик не режутся по периоду (как до изменений)."""
    normalized = (board_code or "").strip().lower()
    return normalized in {DIGITAL_BOARD_CODE, BERCUT_BOARD_CODE}


def _uses_start_date_period(metric: str | None, board_code: str | None) -> bool:
    """Период по start_date / created_at для списка «Всего задач» и досок с датой в метриках."""
    if not metric:
        return True
    if _board_metrics_ignore_date_period(board_code):
        return metric not in {"in_progress", "launching_soon", "launched", "completed", "errors"}
    return True


def _planned_date_upcoming_sort_key(task: Task, *, today: date | None = None) -> tuple[int, date, int]:
    """Будущие — от ближайших; затем прошедшие; TBD и без даты — в конце."""
    today_value = today or date.today()
    planned, quarter_key, _, planned_label = _task_plan_meta(task)
    try:
        number_key = int(task.external_id)
    except ValueError:
        number_key = 0

    if quarter_key == PLAN_QUARTER_TBD or planned_label == "TBD":
        return (2, date.max, number_key)

    if planned is None:
        return (3, date.max, number_key)

    if planned < today_value:
        return (1, date.max - planned, number_key)

    return (0, planned, number_key)


def _title_sort_key(task: Task) -> str:
    return task.title.casefold()


def _sort_key(task: Task, sort: str):
    if sort == "title":
        return _title_sort_key(task)
    if sort == "release_date":
        return task.release_date or date.min
    if sort == "planned_date":
        planned, quarter_key, _, _ = _task_plan_meta(task)
        if quarter_key == PLAN_QUARTER_TBD:
            return date(1900, 1, 1)
        return planned or date.min
    if sort == "start_date":
        return _effective_start(task) or date.min
    if sort == "created_at":
        return task.created_at or date.min
    try:
        return int(task.external_id)
    except ValueError:
        return 0


def _board_name_by_code(code: str | None) -> str | None:
    if not code:
        return None
    for board in get_boards():
        if board.code.lower() == code.lower():
            return board.display_name
    return None


def _collect_available_quarters(rows: list[Task]) -> list[QuarterOptionOut]:
    current_year = date.today().year
    keys: set[str] = set()
    for row in rows:
        _, quarter_key, _, _ = _task_plan_meta(row)
        if not quarter_key or quarter_key == PLAN_QUARTER_TBD:
            continue
        if quarter_key.startswith(f"{current_year}-Q"):
            keys.add(quarter_key)
    return [
        QuarterOptionOut(key=key, label=quarter_label_from_key(key))
        for key in sorted(keys, reverse=True)
    ]


def _closed_states_lower() -> set[str]:
    return {value.lower() for value in settings.closed_state_list}


def _is_closed_zni(task: Task) -> bool:
    return bool(task_status_tokens(task) & _closed_states_lower())


def _build_errors_by_parent(zni_rows: list[Task], error_rows: list[Task]) -> dict[int, list[Task]]:
    zni_id_by_external = {row.external_id: row.id for row in zni_rows}
    zni_db_ids = {row.id for row in zni_rows}
    errors_by_parent: dict[int, list[Task]] = {}
    for error in error_rows:
        parent_id = error.parent_task_id
        if parent_id is None or parent_id not in zni_db_ids:
            raw_parent = _extra(error).get("parent_zni_id")
            if raw_parent is not None:
                parent_id = zni_id_by_external.get(str(raw_parent))
        if parent_id is not None:
            errors_by_parent.setdefault(parent_id, []).append(error)
    return errors_by_parent


def _matches_closed_table_visibility(task: Task, metric: str | None) -> bool:
    """Closed ЗНИ: в таблице при «Завершённые» и «Запущено» (BE: Closed = запущено)."""
    if not _is_closed_zni(task):
        return True
    return metric in {"completed", "launched"}


def _collect_available_statuses(rows: list[Task]) -> list[str]:
    closed_states = _closed_states_lower()
    values: set[str] = set()
    for row in rows:
        column = _board_column(row)
        if column and column.casefold() not in closed_states:
            values.add(column)
        status = row.source_status
        if status and status.casefold() not in closed_states:
            values.add(status)
    return sorted(values, key=str.casefold)


def _matches_metric_filter(
    row: Task,
    metric: str | None,
    *,
    errors_by_parent: dict[int, list[Task]],
    date_from: date | None,
    date_to: date | None,
) -> bool:
    if not metric or metric == "all":
        return True
    today = date.today()
    horizon = today + timedelta(days=settings.launching_soon_days)
    if metric == "in_progress":
        return is_in_progress(row)
    if metric == "launching_soon":
        return is_launching_soon(row, today=today, horizon=horizon)
    if metric == "launched":
        return is_launched(row, date_from=date_from, date_to=date_to)
    if metric == "completed":
        return is_completed(row, date_from=date_from, date_to=date_to)
    if metric == "errors":
        return has_linked_errors(row, errors_by_parent)
    return True


def _matches_dashboard_row(
    row: Task,
    metric: str | None,
    *,
    board_code: str | None,
    errors_by_parent: dict[int, list[Task]],
    date_from: date | None,
    date_to: date | None,
) -> bool:
    row_board_code = _row_board_code(row, board_code)
    if _uses_start_date_period(metric, row_board_code) and not _in_date_range(row, date_from, date_to):
        return False
    if not metric:
        return True
    return _matches_metric_filter(
        row,
        metric,
        errors_by_parent=errors_by_parent,
        date_from=date_from,
        date_to=date_to,
    )


def _compute_metrics(
    rows: list[Task],
    *,
    board_code: str | None,
    error_rows: list[Task],
    errors_by_parent: dict[int, list[Task]],
    date_from: date | None,
    date_to: date | None,
) -> DashboardMetricsOut:
    def _metric_row(row: Task, metric: str) -> bool:
        return _matches_dashboard_row(
            row,
            metric,
            board_code=board_code,
            errors_by_parent=errors_by_parent,
            date_from=date_from,
            date_to=date_to,
        )

    standalone_incident_errors = _standalone_incident_errors(error_rows)
    incident_errors_count = sum(
        1
        for error in standalone_incident_errors
        if _matches_incident_error_row(
            error,
            board_code=board_code,
            metric="errors",
            search=None,
            status=None,
            date_from=date_from,
            date_to=date_to,
        )
    )

    return DashboardMetricsOut(
        totalTasks=sum(1 for row in rows if not _is_closed_zni(row)),
        inProgress=sum(1 for row in rows if _metric_row(row, "in_progress")),
        launchingSoon=sum(1 for row in rows if _metric_row(row, "launching_soon")),
        launched=sum(1 for row in rows if _metric_row(row, "launched")),
        completed=sum(1 for row in rows if _metric_row(row, "completed")),
        errorsCount=sum(1 for row in rows if _metric_row(row, "errors")) + incident_errors_count,
    )


def load_change_requests(
    db: Session,
    *,
    board_code: str | None,
    search: str | None = None,
    sort: str = "id_desc",
    date_from: date | None = None,
    date_to: date | None = None,
    status: str | None = None,
    quarter: str | None = None,
    ect_reservation: str | None = None,
    linked_environment: str | None = None,
    metric: str | None = None,
    tag_groups: list[str] | None = None,
) -> DashboardOut:
    boards = ensure_boards_loaded(db, refresh=True)
    all_boards = is_all_boards(board_code)
    matched = [] if all_boards else boards_for_sync(board_code, boards)
    board = matched[0] if matched else None
    selected_tag_groups = (
        normalize_tag_group_keys(tag_groups, board_code)
        if tag_filter_supported_for_board(board_code)
        else []
    )

    if all_boards:
        scope = _boards_task_clause(boards)
    elif board is None:
        return DashboardOut(
            board=None,
            allBoards=False,
            metrics=_empty_metrics(),
            items=[],
            totalShown=0,
            actualPeriodEditableStatuses=actual_period_editable_statuses(),
        )
    else:
        scope = _boards_task_clause(matched)

    zni_query = select(Task).where(
        Task.task_type == "change_request",
        scope,
    )
    error_query = select(Task).where(
        Task.task_type == "error",
        scope,
    )

    rows = list(db.scalars(zni_query))
    error_rows = active_errors(list(db.scalars(error_query)))

    rows_with_customer = [row for row in rows if has_customer_name(row)]

    errors_by_parent = _build_errors_by_parent(rows, error_rows)

    filtered = [
        row
        for row in rows_with_customer
        if _matches_search(row, search or "")
        and _matches_status(row, status)
        and _matches_quarter(row, quarter)
        and _matches_ect_reservation(row, ect_reservation)
        and _matches_linked_environment(row, linked_environment)
        and _matches_dashboard_row(
            row,
            metric,
            board_code=board_code,
            errors_by_parent=errors_by_parent,
            date_from=date_from,
            date_to=date_to,
        )
        and _matches_tag_groups(row, selected_tag_groups)
        and _matches_closed_table_visibility(row, metric)
    ]

    filtered_incident_errors = [
        error
        for error in _standalone_incident_errors(error_rows)
        if _matches_incident_error_row(
            error,
            board_code=board_code,
            metric=metric,
            search=search,
            status=status,
            date_from=date_from,
            date_to=date_to,
        )
    ]

    combined_rows: list[tuple[str, Task]] = [("change_request", row) for row in filtered]
    combined_rows.extend(("error", error) for error in filtered_incident_errors)

    if sort == "planned_date_upcoming":
        combined_rows.sort(key=lambda pair: _planned_date_upcoming_sort_key(pair[1]))
    elif sort == "business_value_asc":
        combined_rows.sort(key=lambda pair: _business_value_asc_sort_key(pair[1]))
    elif sort == "business_value_desc":
        combined_rows.sort(key=lambda pair: _business_value_desc_sort_key(pair[1]))
    else:
        reverse = sort.endswith("_desc") or sort == "id_desc"
        sort_field = sort.replace("_desc", "").replace("_asc", "")
        combined_rows.sort(key=lambda pair: _sort_key(pair[1], sort_field), reverse=reverse)

    items: list[ChangeRequestOut] = []
    change_request_rows = [row for row_type, row in combined_rows if row_type == "change_request"]
    external_by_id = load_external_data_by_task_ids(
        db,
        [row.id for row in change_request_rows],
    )
    products_by_id = _load_products_by_id(db, change_request_rows)
    for row_type, row in combined_rows:
        if row_type == "error":
            items.append(_incident_error_to_item(row))
            continue
        items.append(
            _change_request_to_out(
                row,
                errors_by_parent.get(row.id, []),
                external_by_id.get(row.id),
                products_by_id.get(row.parent_task_id) if row.parent_task_id else None,
            )
        )

    return DashboardOut(
        board=_board_out(board, boards) if board and not all_boards else None,
        allBoards=all_boards,
        metrics=_compute_metrics(
            rows_with_customer,
            board_code=board_code,
            error_rows=error_rows,
            errors_by_parent=errors_by_parent,
            date_from=date_from,
            date_to=date_to,
        ),
        items=items,
        totalShown=len(items),
        availableStatuses=_collect_available_statuses(
            rows_with_customer + (
                _standalone_incident_errors(error_rows)
                if is_all_boards(board_code)
                or (board_code or "").strip().lower() == BERCUT_BOARD_CODE
                else []
            )
        ),
        availableQuarters=_collect_available_quarters(rows_with_customer),
        availableTagGroups=(
            _tag_filter_groups_out(board_code)
            if tag_filter_supported_for_board(board_code)
            else []
        ),
        actualPeriodEditableStatuses=actual_period_editable_statuses(),
    )


def _boards_for_export(
    board_code: str | None,
    boards: list[BoardConfig] | None = None,
) -> list[BoardConfig]:
    source = boards if boards is not None else get_boards()
    if is_all_boards(board_code) or not board_code:
        return grouped_boards_for_ui(source)
    matched = boards_for_sync(board_code, source)
    return [matched[0]] if matched else []


XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_EXPORT_HEADERS = [
    "Номер ЗНИ",
    "ЗНИ",
    "Заказчик",
    "Продукт",
    "Статус workflow",
    "Статус доски",
    "Дата начала",
    "Целевая дата",
    "Планируемая дата",
    "План квартала",
    "Плановый релиз",
    "Бронь ресурса ЕЦТ",
    "Доска",
    "Ошибки",
    "Приоритет",
    "Категория",
    "Коммерческий эффект",
    "Фактическая дата месяц/квартал",
    "Желаемая дата",
    "Комментарий",
]
_EXPORT_HEADER_FILL = PatternFill(fill_type="solid", fgColor="CCCCCC")
_EXPORT_HEADER_FONT = Font(bold=True)
_EXPORT_LINK_FONT = Font(color="0563C1", underline="single")
_EXPORT_TEXT_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
_EXPORT_PRODUCT_COLUMN = _EXPORT_HEADERS.index("Продукт") + 1
_MOSCOW_TZ = ZoneInfo("Europe/Moscow")


def _product_export_label(item: ChangeRequestOut) -> str:
    product = item.product
    if product is None:
        return ""
    title = (product.title or "").strip()
    if title:
        return f"{product.id}: {title}"
    return product.id


def _export_row_values(item: ChangeRequestOut) -> list[object]:
    errors_text = "; ".join(f"{e.id}: {e.title}" for e in item.errors)
    return [
        item.number,
        item.title,
        item.customerName or "",
        _product_export_label(item),
        item.status or "",
        item.boardColumn or "",
        item.startDate.isoformat() if item.startDate else "",
        item.releaseDate.isoformat() if item.releaseDate else "",
        item.plannedLabel or (item.plannedDate.isoformat() if item.plannedDate else ""),
        item.planQuarter or "",
        item.plannedRelease or "",
        ect_resource_reservation_label(item.ectResourceReservation),
        item.boardName or "",
        errors_text,
        item.externalPriority or "",
        item.externalCategoryName or "",
        item.externalCommercialEffect or "",
        item.externalActualPeriod or "",
        item.externalDesiredDate.isoformat() if item.externalDesiredDate else "",
        item.externalComment or "",
    ]


def _iter_export_items(
    db: Session,
    board_code: str | None,
    boards: list[BoardConfig],
) -> list[ChangeRequestOut]:
    items: list[ChangeRequestOut] = []
    for board in _boards_for_export(board_code, boards):
        single = load_change_requests(db, board_code=board.code)
        for item in single.items:
            if is_excluded_zni_title(item.title):
                continue
            items.append(item)
    return items


def export_xlsx(db: Session, *, board_code: str | None = None) -> tuple[bytes, str]:
    boards = ensure_boards_loaded(db, refresh=True)
    items = _iter_export_items(db, board_code, boards)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "ЗНИ"
    for col_index, title in enumerate(_EXPORT_HEADERS, start=1):
        cell = sheet.cell(row=1, column=col_index, value=title)
        cell.fill = _EXPORT_HEADER_FILL
        cell.font = _EXPORT_HEADER_FONT
        cell.alignment = _EXPORT_TEXT_ALIGNMENT

    for row_index, item in enumerate(items, start=2):
        for col_index, value in enumerate(_export_row_values(item), start=1):
            cell = sheet.cell(row=row_index, column=col_index, value=value)
            cell.alignment = _EXPORT_TEXT_ALIGNMENT
        if item.product and item.product.url:
            product_cell = sheet.cell(row=row_index, column=_EXPORT_PRODUCT_COLUMN)
            product_cell.hyperlink = item.product.url
            product_cell.font = _EXPORT_LINK_FONT

    last_row = max(1, len(items) + 1)
    for col_index in range(1, len(_EXPORT_HEADERS) + 1):
        letter = get_column_letter(col_index)
        max_len = 12
        for row_index in range(1, last_row + 1):
            value = sheet.cell(row=row_index, column=col_index).value
            if value is None:
                continue
            lines = str(value).splitlines() or [""]
            max_len = max(max_len, max(len(line) for line in lines))
        sheet.column_dimensions[letter].width = min(48, max(12, max_len + 2))
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(_EXPORT_HEADERS))}{last_row}"

    buffer = io.BytesIO()
    workbook.save(buffer)
    stamp = datetime.now(_MOSCOW_TZ).strftime("%Y%m%d")
    suffix = "all"
    if board_code and not is_all_boards(board_code):
        suffix = board_code.strip().lower()
    return buffer.getvalue(), f"zni-report-{suffix}-{stamp}.xlsx"
