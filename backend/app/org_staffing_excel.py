"""Excel-экспорт Staffing: общий список сотрудников и листы по отделам."""

from __future__ import annotations

import io
import re
from datetime import datetime
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.org_schemas import DepartmentOut, EmployeeOut

MOSCOW_TZ = ZoneInfo("Europe/Moscow")
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="CCCCCC")
_HEADER_FONT = Font(bold=True)
_TEXT_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
_INVALID_SHEET_CHARS = re.compile(r"[\:\\/?*\[\]]+")
_MAX_SHEET_NAME_LEN = 31

GENERAL_HEADERS = [
    "ФИО",
    "Должность",
    "Отделы",
    "Email",
    "Рабочих часов в день",
    "Экспертиза",
    "Руководитель",
    "Активен",
]

DEPARTMENT_HEADERS = [
    "ФИО",
    "Должность",
    "Email",
    "Рабочих часов в день",
    "Экспертиза",
    "Руководитель",
    "Активен",
]


def _yes_no(value: bool) -> str:
    return "Да" if value else "Нет"


def _format_expertises(employee: EmployeeOut) -> str:
    if not employee.expertises:
        return ""
    parts: list[str] = []
    for item in employee.expertises:
        if item.level:
            parts.append(f"{item.directionName} ({item.level})")
        else:
            parts.append(item.directionName)
    return ", ".join(parts)


def _format_departments(employee: EmployeeOut) -> str:
    if not employee.departments:
        return ""
    return ", ".join(item.departmentName for item in employee.departments if item.departmentName)


def _unique_sheet_name(raw: str, used: set[str]) -> str:
    cleaned = _INVALID_SHEET_CHARS.sub(" ", (raw or "Лист").strip()) or "Лист"
    base = cleaned[:_MAX_SHEET_NAME_LEN].strip() or "Лист"
    candidate = base
    suffix = 1
    while candidate.casefold() in used:
        tail = f" {suffix}"
        candidate = f"{base[: _MAX_SHEET_NAME_LEN - len(tail)]}{tail}".strip()
        suffix += 1
    used.add(candidate.casefold())
    return candidate


def _autosize_columns(worksheet: Worksheet, *, column_count: int, row_count: int) -> None:
    for index in range(1, column_count + 1):
        letter = get_column_letter(index)
        max_len = 12
        for row_index in range(1, row_count + 1):
            value = worksheet.cell(row=row_index, column=index).value
            if value is None:
                continue
            lines = str(value).splitlines() or [""]
            max_len = max(max_len, max(len(line) for line in lines))
        worksheet.column_dimensions[letter].width = min(48, max(12, max_len + 2))


def _write_header(worksheet: Worksheet, headers: list[str]) -> None:
    for col_index, title in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=col_index, value=title)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _TEXT_ALIGNMENT


def _write_general_row(worksheet: Worksheet, row_index: int, employee: EmployeeOut) -> None:
    values = [
        employee.fullName,
        employee.position or "",
        _format_departments(employee),
        employee.email or "",
        float(employee.dailyWorkHours),
        _format_expertises(employee),
        employee.managerName or "",
        _yes_no(employee.isActive),
    ]
    for col_index, value in enumerate(values, start=1):
        cell = worksheet.cell(row=row_index, column=col_index, value=value)
        cell.alignment = _TEXT_ALIGNMENT


def _write_department_row(worksheet: Worksheet, row_index: int, employee: EmployeeOut) -> None:
    values = [
        employee.fullName,
        employee.position or "",
        employee.email or "",
        float(employee.dailyWorkHours),
        _format_expertises(employee),
        employee.managerName or "",
        _yes_no(employee.isActive),
    ]
    for col_index, value in enumerate(values, start=1):
        cell = worksheet.cell(row=row_index, column=col_index, value=value)
        cell.alignment = _TEXT_ALIGNMENT


def _finish_sheet(worksheet: Worksheet, *, headers: list[str], data_rows: int) -> None:
    last_row = max(1, data_rows + 1)
    _autosize_columns(worksheet, column_count=len(headers), row_count=last_row)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last_row}"


def _visible_in_pyramid(employee: EmployeeOut) -> bool:
    return bool(employee.isActive) and not bool(employee.hideFromPyramid)


def generate_staffing_excel(
    employees: list[EmployeeOut],
    departments: list[DepartmentOut],
) -> tuple[bytes, str]:
    """
    Лист 1 — все, кто попадает в пирамиду (активен и без «Не отображать в пирамиде»).
    Далее — по одному листу на отдел, если в нём есть такие сотрудники.
    """
    visible = [emp for emp in employees if _visible_in_pyramid(emp)]
    visible.sort(key=lambda emp: emp.fullName.casefold())

    by_department: dict[int, list[EmployeeOut]] = {}
    for emp in visible:
        for membership in emp.departments:
            by_department.setdefault(membership.departmentId, []).append(emp)

    for rows in by_department.values():
        rows.sort(key=lambda emp: emp.fullName.casefold())

    workbook = Workbook()
    used_names: set[str] = set()

    general = workbook.active
    general.title = _unique_sheet_name("Общий список", used_names)
    _write_header(general, GENERAL_HEADERS)
    for row_index, emp in enumerate(visible, start=2):
        _write_general_row(general, row_index, emp)
    _finish_sheet(general, headers=GENERAL_HEADERS, data_rows=len(visible))

    ordered_departments = sorted(
        departments,
        key=lambda dept: (dept.sortOrder, dept.name.casefold(), dept.id),
    )
    for dept in ordered_departments:
        members = by_department.get(dept.id) or []
        if not members:
            continue
        sheet = workbook.create_sheet(title=_unique_sheet_name(dept.name, used_names))
        _write_header(sheet, DEPARTMENT_HEADERS)
        for row_index, emp in enumerate(members, start=2):
            _write_department_row(sheet, row_index, emp)
        _finish_sheet(sheet, headers=DEPARTMENT_HEADERS, data_rows=len(members))

    buffer = io.BytesIO()
    workbook.save(buffer)
    stamp = datetime.now(MOSCOW_TZ).strftime("%Y%m%d")
    filename = f"staffing-employees-{stamp}.xlsx"
    return buffer.getvalue(), filename
