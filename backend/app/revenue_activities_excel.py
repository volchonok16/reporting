"""Excel-экспорт «Активности по выручкам»: числа как number, текст как строка."""

from __future__ import annotations

import io
from datetime import datetime
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.product_status_rich_text import display_cell_text
from app.revenue_activities_db import (
    REVENUE_ACTIVITY_SECTION_COLUMNS,
    REVENUE_NUMERIC_COLUMNS,
    _parse_numeric,
)
from app.schemas import ProductStatusB2BOut, ProductStatusSheetOut

MOSCOW_TZ = ZoneInfo("Europe/Moscow")
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="CCCCCC")
_HEADER_FONT = Font(bold=True)
_TEXT_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
_NUMBER_ALIGNMENT = Alignment(horizontal="right", vertical="top")
_NUMBER_FORMAT = "0.####"
_TOTALS_FILL = PatternFill(fill_type="solid", fgColor="E8E8E8")
_TOTALS_FONT = Font(bold=True)


def _is_numeric_export_column(column: str) -> bool:
    return column in REVENUE_NUMERIC_COLUMNS


def _cell_export_value(column: str, raw: str) -> str | float | None:
    text = display_cell_text(raw).strip()
    if not text:
        return None
    if _is_numeric_export_column(column):
        parsed = _parse_numeric(text)
        if parsed is not None:
            return parsed
        return text
    return text


def _autosize_columns(
    worksheet: Worksheet,
    *,
    column_count: int,
    row_count: int,
) -> None:
    for index in range(1, column_count + 1):
        letter = get_column_letter(index)
        max_len = 12
        for row_index in range(1, row_count + 1):
            value = worksheet.cell(row=row_index, column=index).value
            if value is None:
                continue
            lines = str(value).splitlines() or [""]
            max_len = max(max_len, max(len(line) for line in lines))
        width = min(48, max(12, max_len + 2))
        worksheet.column_dimensions[letter].width = width


def _column_totals(sheet: ProductStatusSheetOut, columns: list[str]) -> dict[str, float | None]:
    totals: dict[str, float | None] = {}
    for column in columns:
        if not _is_numeric_export_column(column):
            totals[column] = None
            continue
        total = 0.0
        has_value = False
        for row in sheet.rows:
            raw = row.get(column, "") if column else ""
            text = display_cell_text(raw if isinstance(raw, str) else str(raw or "")).strip()
            parsed = _parse_numeric(text)
            if parsed is None:
                continue
            total += parsed
            has_value = True
        totals[column] = total if has_value else None
    return totals


def _write_revenue_sheet(worksheet: Worksheet, sheet: ProductStatusSheetOut) -> None:
    columns = list(sheet.columns) or list(REVENUE_ACTIVITY_SECTION_COLUMNS["main"])
    for col_index, column_name in enumerate(columns, start=1):
        cell = worksheet.cell(row=1, column=col_index, value=column_name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _TEXT_ALIGNMENT

    for row_index, row in enumerate(sheet.rows, start=2):
        for col_index, column_name in enumerate(columns, start=1):
            raw = row.get(column_name, "") if column_name else ""
            value = _cell_export_value(column_name, raw if isinstance(raw, str) else str(raw or ""))
            cell = worksheet.cell(row=row_index, column=col_index, value=value)
            if _is_numeric_export_column(column_name) and isinstance(value, (int, float)):
                cell.alignment = _NUMBER_ALIGNMENT
                cell.number_format = _NUMBER_FORMAT
            else:
                cell.alignment = _TEXT_ALIGNMENT

    totals_row_index = len(sheet.rows) + 2
    column_totals = _column_totals(sheet, columns)
    for col_index, column_name in enumerate(columns, start=1):
        if col_index == 1:
            cell = worksheet.cell(row=totals_row_index, column=col_index, value="Итого")
            cell.alignment = _TEXT_ALIGNMENT
        else:
            total = column_totals.get(column_name)
            cell = worksheet.cell(
                row=totals_row_index,
                column=col_index,
                value=total if total is not None else None,
            )
            if total is not None:
                cell.alignment = _NUMBER_ALIGNMENT
                cell.number_format = _NUMBER_FORMAT
            else:
                cell.alignment = _TEXT_ALIGNMENT
        cell.fill = _TOTALS_FILL
        cell.font = _TOTALS_FONT

    _autosize_columns(
        worksheet,
        column_count=len(columns),
        row_count=max(1, totals_row_index),
    )
    worksheet.freeze_panes = "A2"
    if sheet.rows:
        last_data_row = len(sheet.rows) + 1
        worksheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{last_data_row}"
    else:
        worksheet.auto_filter.ref = worksheet.dimensions


def generate_revenue_activities_excel(data: ProductStatusB2BOut) -> tuple[bytes, str]:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    sheets = data.sheets or []
    if not sheets:
        worksheet = workbook.create_sheet(title="Активности по выручкам")
        _write_revenue_sheet(
            worksheet,
            ProductStatusSheetOut(
                gid="main",
                name="Активности по выручкам",
                columns=list(REVENUE_ACTIVITY_SECTION_COLUMNS["main"]),
                rows=[],
                totalShown=0,
            ),
        )
    else:
        for sheet in sheets:
            title = (sheet.name or "Активности по выручкам")[:31]
            worksheet = workbook.create_sheet(title=title or "Лист")
            _write_revenue_sheet(worksheet, sheet)

    buffer = io.BytesIO()
    workbook.save(buffer)
    stamp = datetime.now(MOSCOW_TZ).strftime("%Y%m%d")
    filename = f"aktivnosti-po-vyruchkam-{stamp}.xlsx"
    return buffer.getvalue(), filename
